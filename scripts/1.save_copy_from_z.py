# -*- coding: utf-8 -*-
# PYTHON 3
#----------------------------------------------------------------------------
# Created By  : Mélanie Aubry - aubry.melanie33@gmail.com
# Created Date: March 29, 2022
# ---------------------------------------------------------------------------

import openpyxl as xl;
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from definitions import INPUT_DIR

wb2 = Workbook()
path2 =f'{INPUT_DIR}/global_save.xlsx'
wb2.save(path2)

path1 ="Z:/SCD/Pôle Services Numériques/Numérisation/Cartothèque/aubry_melanie_2022_etat_des_lieux.xlsx"
wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

global_save = f'{INPUT_DIR}/global_save.xlsx'
wb2 = xl.load_workbook(global_save)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

del wb2['Sheet']

wb2.save(global_save)

df = pd.read_excel(f'{INPUT_DIR}/global_save.xlsx')
df.to_csv(f'{INPUT_DIR}/global_save.csv')

print("COPIED !")