# -*- coding: utf-8 -*-
# PYTHON 3
#----------------------------------------------------------------------------
# Created By  : Mélanie Aubry - aubry.melanie33@gmail.com
# Created Date: May 11, 2022
# ---------------------------------------------------------------------------

from definitions import OUTPUT_DIR
from utils.NakalaUtils import post_relations
import openpyxl

HANDLE_COLUMN_INDEXES = [2,3,4,5,6]

workbook = openpyxl.load_workbook(f'{OUTPUT_DIR}/test_relations.xlsx')
#workbook = openpyxl.load_workbook(f'{OUTPUT_DIR}/relations__20220517.xlsx')
worksheet = workbook.get_sheet_by_name('Feuil1')

handles_by_relation = {}
for index, row_cells in enumerate(worksheet.iter_rows()):
    if index == 0 or row_cells[7].value is None:
        continue
    relation = row_cells[7].value
    if not relation in handles_by_relation:
        handles_by_relation[relation] = []
    for handle_column_index in HANDLE_COLUMN_INDEXES:
        handle_nakala = row_cells[handle_column_index].value
        if handle_nakala is not None:
            handles_by_relation[relation].append(handle_nakala)

for relation in handles_by_relation:
    if len(handles_by_relation[relation]) < 2:
        continue
    for handle_nakala in handles_by_relation[relation]:
        relations = []
        for other_handle_nakala in handles_by_relation[relation]:
            if handle_nakala == other_handle_nakala:
                continue
            relations.append({
                "type": "References",
                "repository": "nakala",
                "target": other_handle_nakala,
                "comment": ""
            })
        post_relations(handle_nakala, relations)