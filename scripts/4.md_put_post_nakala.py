# -*- coding: utf-8 -*-
# PYTHON 3
#----------------------------------------------------------------------------
# Created By  : Mélanie Aubry - aubry.melanie33@gmail.com
# Created Date: April 25, 2022
# ---------------------------------------------------------------------------

from cmath import nan
import json
from definitions import COLUMN_NAME_BY_META_PROPERTY_NAKALA, TYPE_BY_META_PROPERTY_NAKALA, META_PROPERTY_NAKALA_BY_COLUMN_NAME, OUTPUT_DIR, DATA_PROPERTY_BY_COLUMN_INDEX
from utils.ArrayUtils import try_removing_value_in_array
from utils.NakalaUtils import get_data, post_metadata, put_data
from utils.StringUtils import equal_ignore_cases
import openpyxl

NAKALA_HANDLE_COLUMN_INDEXES = [13, 14, 15, 16, 17, 18]

class NakalaService:
    nakala_metadatas_to_override = ['title', 'date', 'type', 'creator', 'license']
    nakala_metadatas_to_add = ['ppn', 'f', 'subject', 'language', 'spatial', 'extent', 'publisher', 'description', 'cote']
    
    def __init__(self, data, nakala_id):
        self.data = data
        self.nakala_id = nakala_id

    def upload_data_into_nakala(self):
        self.nakala_data = get_data(self.nakala_id)
        data_needs_to_be_updated = self.check_metadatas_for_change()
        if data_needs_to_be_updated:
            print('EDITING NAKALA DATA BY API...')
            print(f'NAKALA DATA BEFORE API CALLS : {json.dumps(self.nakala_data, indent=4)}')
            if self.nakala_metadatas_to_override:
                self.put_new_metadatas_in_nakala()
            if self.nakala_metadatas_to_add:
                self.post_new_metadatas_in_nakala()
            print(f'NAKALA DATA AFTER API CALLS : {json.dumps(get_data(self.nakala_id), indent=4)}')
        else:
            print('NO NEED TO EDIT METADATAS')
    
    def check_metadatas_for_change(self):
        print('CHECKING FOR CHANGE BETWEEN NAKALA DATA AND FILE')
        for meta in self.nakala_data['metas']:
            if self.metadata_needs_be_updated(meta):
                continue
            self.remove_column_from_metadata_to_put_or_post(meta)
        return self.nakala_metadatas_to_override or self.nakala_metadatas_to_add

    def metadata_needs_be_updated(self, meta):
        return not meta['propertyUri'] in COLUMN_NAME_BY_META_PROPERTY_NAKALA or meta['value'] is None or equal_ignore_cases(meta['value'], 'a definir')

    def remove_column_from_metadata_to_put_or_post(self, meta):
        propertyUri = meta['propertyUri']
        if propertyUri in COLUMN_NAME_BY_META_PROPERTY_NAKALA:
            column_name = COLUMN_NAME_BY_META_PROPERTY_NAKALA[propertyUri]
            try_removing_value_in_array(self.nakala_metadatas_to_override, column_name)
            try_removing_value_in_array(self.nakala_metadatas_to_add, column_name)

    def put_new_metadatas_in_nakala(self):
        for meta in self.nakala_data['metas']:
            propertyUri = meta['propertyUri']
            if propertyUri in COLUMN_NAME_BY_META_PROPERTY_NAKALA:
                column_name = COLUMN_NAME_BY_META_PROPERTY_NAKALA[propertyUri]
                if column_name in self.nakala_metadatas_to_override:
                    if propertyUri in TYPE_BY_META_PROPERTY_NAKALA:
                        meta['typeUri'] = TYPE_BY_META_PROPERTY_NAKALA[propertyUri]
                    meta['value'] = self.retrieve_data_value(column_name)
        put_data(self.nakala_data)
        
    def retrieve_data_value(self, column_name):
        if column_name == 'creator':
            if self.data['creatorfn'] is None and self.data['creatorln'] is None:
                return { 'surname': self.data['creatorentity'] }
            return {
                'givenname': self.data['creatorfn'], 
                'surname': self.data['creatorln']
            }
        return self.data[column_name]

    def post_new_metadatas_in_nakala(self):
        print('ADDING NEW METADATAS')
        for column_name in self.nakala_metadatas_to_add:
            if column_name in self.data:
                if column_name == 'subject':
                    for keyword in self.data['subject'].split('#'):
                        self.post_metadata_in_nakala(column_name, keyword)
                else:
                    value = self.retrieve_data_value(column_name)
                    self.post_metadata_in_nakala(column_name, value)

    def post_metadata_in_nakala(self, column_name, value):
        propertyUri = META_PROPERTY_NAKALA_BY_COLUMN_NAME[column_name]
        meta_to_post = {
            "value": value,
            "propertyUri": propertyUri
        }
        if propertyUri in TYPE_BY_META_PROPERTY_NAKALA:
            meta_to_post['typeUri'] = TYPE_BY_META_PROPERTY_NAKALA[propertyUri]
        post_metadata(self.nakala_id, meta_to_post)

workbook = openpyxl.load_workbook(f'{OUTPUT_DIR}/testnakala.xlsx')
#workbook = openpyxl.load_workbook(f'{OUTPUT_DIR}/metadata_20220513.xlsx')
worksheet = workbook.get_sheet_by_name('Feuil1')
for index, row_cells in enumerate(worksheet.iter_rows()):
    if index == 0:
        continue
    handle_ids = []
    data = {}
    for column_index in DATA_PROPERTY_BY_COLUMN_INDEX:
        if row_cells[column_index].value is not None:
            property = DATA_PROPERTY_BY_COLUMN_INDEX[column_index]
            data[property] = row_cells[column_index].value
    for column_index in NAKALA_HANDLE_COLUMN_INDEXES:
        handle_nakala = row_cells[column_index].value
        if handle_nakala is not None:
            NakalaService(data, handle_nakala).upload_data_into_nakala()